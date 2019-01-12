import openpyxl

def getmsg():
    '''
    :return:dic，索引是第一列，数据元素是第2 3 4列构成的列表
    '''
    dic={}
    a=openpyxl.load_workbook("1.xlsx")
    print(a.sheetnames)
    sheet1 = a["Sheet1"]
    for i in range(2,52):
        list=[]
        list.append(sheet1["B%d" % i].value)
        list.append(sheet1["C%d" % i].value)
        list.append(sheet1["D%d" % i].value)
        dic[int(sheet1["A%d" % i].value)]=list
    return dic

def deal():
    a=openpyxl.load_workbook("3.xlsx")
    sheet1 = a["Sheet1"]
    msg=getmsg()
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(1,644):
        if(int(sheet1["A%d" % i].value) in msg.keys()):
            print(int(sheet1["A%d" % i].value))
            print(msg[int(sheet1["A%d" % i].value)][0])
            msg[int(sheet1["A%d" % i].value)].append(sheet1["A%d" % i].value)
            ws.append(msg[int(sheet1["A%d" % i].value)])
            msg[int(sheet1["A%d" % i].value)].remove(sheet1["A%d" % i].value)
        else:
            list=[]
            list.append(sheet1["A%d" % i].value)
            ws.append(list)
    wb.save('4.xlsx')
    a.close()
deal()